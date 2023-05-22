import json

import openpyxl
import requests
from config import *

schools = list()
exists_schools = list()
school_max_id = 1000


def get_exists_schools():
    url = f"{DOMJUDGE_URL}/api/contests/2/organizations"
    res = requests.get(url, headers={"cookie": COOKIE})
    for school in res.json():
        exists_schools.append({"id": school['id'], "name": school['name']})


def add_school(name):
    global school_max_id
    school = {"icpc_id": f"school{school_max_id}", "name": name, "short_name": name}
    school_max_id += 1
    schools.append(school)
    return school


def find_school(name):
    for school in exists_schools:
        if school["name"] == name:
            return school
    for school in schools:
        if school["name"] == name:
            return school
    return None


def read_excel(file_name):
    wb = openpyxl.load_workbook(file_name)
    ws = wb.worksheets[0]
    for row in range(2, ws.max_row + 1):
        school_name = ws.cell(row, SCHOOL_NAME_COL).value
        school = find_school(school_name)
        if school is None:
            add_school(school_name)


def upload_schools():
    url = f"{DOMJUDGE_URL}/jury/affiliations/add"
    print(f"已存在学校数量:{len(exists_schools)},需添学校数量{len(schools)}")
    for school in schools:
        data = {
            "team_affiliation[icpcid]": school['icpc_id'],
            "team_affiliation[shortname]": school['short_name'],
            "team_affiliation[name]": school['name'],
            "team_affiliation[country]": "CHN",
            "team_affiliation[internalcomments]": "",
            "team_affiliation[logoFile]": "",
            "team_affiliation[save]": "",
        }

        res = requests.post(url, data=data, headers={"cookie": COOKIE})
        if res.status_code != 200:
            print(f"添加学校{school['name']}失败{res.status_code}")


if __name__ == "__main__":
    get_exists_schools()

    file_name = "data/teams.xlsx"
    read_excel(file_name)
    # save_all()
    upload_schools()
