import openpyxl
import requests
from config.config import *

teams = list()
schools = list()

team_max_id = 1000


def get_schools():
    global schools
    url = f"{DOMJUDGE_URL}/api/contests/2/organizations"
    res = requests.get(url, headers={"cookie": COOKIE})
    schools = res.json()


def add_team(name, display_name, school_id, team_type: str, location):
    global team_max_id
    if team_type.find("打星") != -1:
        group = OBSERVERS_ID
    elif team_type.find("女") != -1:
        group = GIRLS_ID
    else:
        group = PARTICIPANTS_ID
    team = {"icpc_id": f"team{team_max_id}", "category": [group, ], "name": name, "display_name": display_name,
            "affiliation": str(school_id), "location": location, "username": f"team{team_max_id}"}
    team_max_id += 1
    teams.append(team)
    return team


def find_school(name):
    for school in schools:
        if school["name"] == name:
            return school
    return None


def read_excel(file_name):
    wb = openpyxl.load_workbook(file_name)
    ws = wb.worksheets[0]
    for row in range(2, ws.max_row + 1):
        team_name = ws.cell(row, TEAM_ENGLISH_NAME_COL).value
        display_name = ws.cell(row, TEAM_CHINESE_NAME_COL).value
        school_name = ws.cell(row, SCHOOL_NAME_COL).value
        team_type = ws.cell(row, TEAM_TYPE_COL).value
        location = ws.cell(row, LOCATION_COL).value
        school = find_school(school_name)
        if school is None:
            raise AssertionError(f"队伍{team_name}所属的学校{school_name}不存在")
        add_team(team_name, display_name, school["id"], team_type, location)


def upload_teams():
    url = f"{DOMJUDGE_URL}/jury/teams/add"
    print(f"总队伍数量:{len(teams)}")
    for team in teams:
        data = {
            "team[icpcid]": team['icpc_id'],
            "team[name]": team['name'],
            "team[displayName]": team['display_name'],
            "team[category]": team['category'],
            "team[publicdescription]": "",
            "team[affiliation]": team['affiliation'],
            "team[penalty]": "0",
            "team[room]": team['location'],
            "team[internalcomments]": "",
            "team[enabled]": "1",
            "team[addUserForTeam]": "create-new-user",
            "team[existingUser]": "1",
            "team[newUsername]": team['username'],
            "team[photoFile]": "",
            "team[save]": "",
        }

        res = requests.post(url, data=data, headers={"cookie": COOKIE})
        if res.status_code != 200:
            print(f"添加队伍{team['name']}失败:{res.status_code}")


if __name__ == "__main__":
    get_schools()
    file_name = "../data/teams.xlsx"
    read_excel(file_name)
    upload_teams()
