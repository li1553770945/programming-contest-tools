import os.path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Side, Border
from config.config import *
import requests
from bs4 import BeautifulSoup

teams = list()


def get_teams():
    global teams
    url = f"{DOMJUDGE_URL}/jury/teams"
    res = requests.get(url, headers={"cookie": COOKIE})
    if res.status_code != 200:
        raise AssertionError(f"访问domjudge获取队伍信息失败:{res.status_code}")
    soup = BeautifulSoup(res.text, 'html.parser')

    # 使用find方法查找具有指定class的表格
    table = soup.select('table[class*="data-table"]')[0]

    # 遍历表格的每一行，并提取数据
    data = []
    for row in table.find_all('tr'):
        row_data = []
        for cell in row.find_all('td'):
            row_data.append(cell.text.strip())
        if row_data:
            data.append(row_data)

    # 打印提取到的数据
    for row in data:
        url = f"{DOMJUDGE_URL}/api/contests/{CONTEST_ID}/teams/{row[0]}"
        res = requests.get(url)
        if res.status_code != 200:
            print(f"队伍id:{row[0]}请求API数据失败:{res.text}")
            continue
        team = res.json()
        teams.append({'id': row[0], 'icpc_id': row[1], 'name': team['name'], 'location': row[7]})


def get_team_index(team_icpc_id):
    for i in range(0, len(teams)):
        if teams[i]['icpc_id'] == team_icpc_id:
            return i
    return -1


def read_tsv():
    with open("data/accounts.tsv", "r", encoding="utf8") as f:
        read_teams = f.readlines()
        first = True
        for read_team in read_teams:
            if first:
                first = False
                continue

            read_team = read_team.replace("\n", "").split("\t")
            team_icpc_id = read_team[2]
            team_password = read_team[3]
            index = get_team_index(team_icpc_id)
            if index == -1:
                print(f"icpc_id为“{team_icpc_id}”的队伍无法从domjudge中找到")
                continue
            teams[index]['password'] = team_password


def save_teams():
    global teams
    workbook = Workbook()

    # 选择默认的活动工作表
    ws = workbook.active

    # 创建一个Alignment对象并将它应用于A1单元格
    align = Alignment(horizontal='center', vertical='center')
    border_style = Side(border_style='thin', color='000000')

    # 创建一个Border对象，指定边框的边界
    border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
    row = 1
    teams = sorted(teams, key=lambda x: x['location'])
    for team in teams:
        if "password" in team:
            for i in range(row, row + 2):
                for j in range(1, 5):
                    ws.cell(i, j).alignment = align
                    ws.cell(i, j).border = border

            ws.cell(row, 1).value = "teamname"
            ws.cell(row, 2).value = "location"
            ws.cell(row, 3).value = "username"
            ws.cell(row, 4).value = "password"
            ws.cell(row + 1, 1).value = team['name']
            ws.cell(row + 1, 2).value = team['location']
            ws.cell(row + 1, 3).value = team['icpc_id']
            ws.cell(row + 1, 4).value = team['password']
            row = row + 4

    if not os.path.exists("../result"):
        os.mkdir("../result")
    workbook.save("../result/accounts.xlsx")


if __name__ == "__main__":
    get_teams()
    read_tsv()
    print(teams[1])
    save_teams()
