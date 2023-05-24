from openpyxl import Workbook

from config.config import *
import requests
from requests.auth import HTTPBasicAuth

teams = list()
schools = list()
results = dict()


def get_teams():
    global teams
    url = f"{DOMJUDGE_URL}/api/v4/contests/{CONTEST_ID}/teams"
    res = requests.get(url, auth=HTTPBasicAuth(ADMIN_USERANME, ADMIN_PASSWORD))
    if res.status_code != 200:
        raise AssertionError(f"API获取队伍失败：{res.status_code}")

    teams = res.json()

def get_schools():
    global schools
    url = f"{DOMJUDGE_URL}/api/v4/contests/{CONTEST_ID}/organizations"
    res = requests.get(url, auth=HTTPBasicAuth(ADMIN_USERANME, ADMIN_PASSWORD))
    if res.status_code != 200:
        raise AssertionError(f"API获取学校失败：{res.status_code}")

    schools = res.json()

def get_team(team_id):
    for i in range(0, len(teams)):
        if teams[i]['id'] == team_id:
            return teams[i]
    return None

def get_school(school_id):
    for i in range(0,len(schools)):
        if schools[i]['id'] == school_id:
            return schools[i]
    return None

def get_results():
    global results
    url = f"{DOMJUDGE_URL}/api/v4/contests/{CONTEST_ID}/scoreboard"
    res = requests.get(url, auth=HTTPBasicAuth(ADMIN_USERANME, ADMIN_PASSWORD))
    if res.status_code != 200:
        raise AssertionError(f"API获取队伍失败：{res.status_code}")

    results = res.json()

def get_team_category_name(category_id):
    if category_id == GIRLS_ID:
        return "女队"
    elif category_id == PARTICIPANTS_ID:
        return "普通"
    elif category_id == OBSERVERS_ID:
        return "打星"
    else:
        raise AssertionError(f"类别id{category_id}不存在，请检查config.py")

def save_results():
    global teams
    workbook = Workbook()

    # 选择默认的活动工作表
    ws = workbook.active
    ws.cell(1, 1).value = "学校"
    ws.cell(1, 2).value = "队伍"
    ws.cell(1, 3).value = "队伍类别"
    ws.cell(1, 4).value = "解题数"
    ws.cell(1, 5).value = "罚时"

    row = 2
    for result in results['rows']:
        team = get_team(result['team_id'])
        if team is None:
            raise AssertionError(f"队伍id{result['team_id']}不存在")

        school = get_school(team['organization_id'])
        if school is None:
            raise AssertionError(f"学校id{team['organization_id']}不存在")

        ws.cell(row,1).value = school['name']
        ws.cell(row,2).value = team['display_name']
        ws.cell(row,3).value = get_team_category_name(team['group_ids'][0])
        ws.cell(row,4).value = result['score']['num_solved']
        ws.cell(row,5).value = result['score']['total_time']
        row += 1

    workbook.save("../result/result.xlsx")


if __name__ == "__main__":
    get_teams()
    get_schools()
    get_results()
    save_results()
